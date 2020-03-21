import os
import shutil

from openpyxl import load_workbook
from starlette.status import HTTP_422_UNPROCESSABLE_ENTITY

from app.core.config import PROJECT_BASE_DIR
from app.exceptions import CustomHTTPException


class FileReaderUtil:
    file_directory = PROJECT_BASE_DIR + '/uploads/'

    def save_file(self, file_object):
        if file_object:
            try:
                file_path = self.file_directory + file_object.filename
                with open(file_path, 'wb+') as buffer:
                    shutil.copyfileobj(file_object.file, buffer)
            finally:
                file_object.file.close()

        return {'path': file_path}

    def remove_file(self, file_path):
        try:
            os.remove(file_path)
        except:
            raise CustomHTTPException(status_code=HTTP_422_UNPROCESSABLE_ENTITY, detail="File does not exists")

        pass

    def excel_to_json_rule_data(self, file_path):
        wb = load_workbook(file_path)

        current_sheet = wb[wb.sheetnames[0]]
        data = []
        for row in current_sheet.values:
            temp_list = [str(x) if x is not None else '' for x in row]
            data.append(temp_list)

        return {'data': data}

